#!/usr/bin/env python3
"""
Build a data-collection reconciliation workbook: report-suite variables
(eVars/props/events/lists) mapped to the property's data elements, showing what
is actually collecting data vs what is not wired up, plus data issues and
prioritized recommendations. See references/reconciliation-workbook-format.md
for the full tab spec.

Usage:
    python3 build_reconciliation_workbook.py input.json output.xlsx

The input JSON is produced by the skill (SKILL.md Phase 2-3) from the audit
workbook + the client's report-suite variable export. It is a NORMALIZED
description of the reconciliation — this script only renders it; it does no
analysis of its own. Shape (all sections optional except reportSuite):

{
  "meta": {
    "property": "AboutAmazon-US (WebSDK)",
    "reportSuite": "aboutamznprod",
    "captured": "2026-07-29",
    "subtitle": "Report suite variables mapped to the property's data elements ...",
    "note": "migration parity confirmed",
    "bottomLine": "Migration is at full parity — ..."     // shown on Overview
  },

  // The report suite's defined variables (from Report Suite Manager). Keys are
  // the number as a string; value is the variable's display name.
  "reportSuite": {
    "eVars":  {"1": "Domain", "5": "Internal Campaign", ...},
    "props":  {"1": "Page Name", ...},
    "events": {"1": "Social Interactions", ...},
    "lists":  {"1": "Errors", ...}                         // optional
  },

  // Which data element(s) + rule(s) feed each variable, and the element's source
  // (data-layer path, "custom code", "s.<field>", etc.). Omit a variable to
  // leave it unpopulated. A variable present here is "Collected" unless one of
  // its elements is in noopElements (then "No-op") or it has a statusOverride.
  "mapping": {
    "eVars":  {"1": [{"element": "Site: Domain", "rule": "Global Page Load Rule",
                      "source": "web.webPageDetails.site"}], ...},
    "props":  { ... },
    "events": {"4": [{"element": "Campaign: Internal Count", "rule": "Global Page Load Rule",
                      "kind": "value"}], ...},             // events: kind "value"/"counter" optional
    "lists":  {"1": [{"element": "Page: Error", "rule": "Global Page Load Rule"},
                     {"element": "Form: Error Name", "rule": "Form Error Tracking"}]}
  },

  // Elements that are wired but return no value (console.log/return "" stubs).
  // Any variable they feed is flagged "No-op (element returns nothing)".
  "noopElements": ["User: Type", "Campaign: Internal"],

  // Force a status + note for specific variables that mapping alone can't infer
  // (ECID excluded by design; audio never built; report-suite duplicates; etc.).
  // status ∈ collected|no-op|not-sent|not-implemented|excluded
  "statusOverrides": {
    "eVars":  {"14": {"status": "excluded", "note": "ECID — Web SDK manages identity natively."},
               "51": {"status": "not-implemented", "note": "AUDIO — defined but never built."}},
    "props":  {"13": {"status": "excluded", "note": "ECID — SDK manages identity."}},
    "events": {"35": {"status": "not-implemented", "note": "AUDIO event — never built."}},
    "lists":  {}
  },

  // Every data element and what it feeds (element → variable view). status is
  // optional; if omitted it is inferred: no-op if in noopElements, orphan if it
  // feeds nothing, else collected.
  "elements": [
    {"name": "Site: Domain", "feeds": ["eVar1 (Domain)"], "status": "collected", "note": ""},
    {"name": "User: Login ID", "feeds": [], "status": "orphan",
     "note": "Reads loginID but no rule sends it. PII review needed."}
  ],

  // Data issues (Data Issues tab). severity ∈ Info|Low|Medium|High
  "issues": [
    {"issue": "User Type never collects", "severity": "Medium",
     "affected": "eVar15 · User: Type",
     "description": "...", "recommendation": "..."}
  ],

  // Prioritized recommendations (Recommendations tab). tier ∈ P1|P2|P3
  "recommendations": [
    {"tier": "P1", "recommendation": "Decide User Type",
     "why": "...", "when": "Before cutover"}
  ]
}
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- status vocabulary + fills (Slalom reconciliation palette) ----
ST_COLLECTED = 'Collected'
ST_NOOP = 'No data (element is a no-op)'
ST_ORPHAN = 'Not wired (orphan element)'
ST_NOTSENT = 'Defined, property never sends it'
ST_NOTIMPL = 'Not implemented (never built)'
ST_EXCLUDED = 'Excluded by design'
CANON = {'collected': ST_COLLECTED, 'no-op': ST_NOOP, 'noop': ST_NOOP, 'orphan': ST_ORPHAN,
         'not-sent': ST_NOTSENT, 'notsent': ST_NOTSENT, 'not-implemented': ST_NOTIMPL,
         'notimpl': ST_NOTIMPL, 'excluded': ST_EXCLUDED}
FILL = {ST_COLLECTED: 'C6EFCE', ST_NOOP: 'FFEB9C', ST_ORPHAN: 'FFE699', ST_NOTSENT: 'FCE4D6',
        ST_NOTIMPL: 'FFC7CE', ST_EXCLUDED: 'D9D9D9'}
FONTCOL = {ST_COLLECTED: '006100', ST_NOOP: '9C5700', ST_ORPHAN: '7F6000', ST_NOTSENT: '843C0C',
           ST_NOTIMPL: '9C0006', ST_EXCLUDED: '808080'}
DESC = {
    ST_COLLECTED: 'Element feeds the variable and a rule sends it — data is collected.',
    ST_NOOP: 'Element is wired but returns no value (no-op) — the variable never populates. Decision needed.',
    ST_ORPHAN: 'Element exists but is not sent to any variable (data-element side only).',
    ST_NOTSENT: 'Defined in the report suite but the property never populates it (stays empty).',
    ST_NOTIMPL: 'Defined in the report suite but never built anywhere — net-new scope if wanted (e.g. Audio).',
    ST_EXCLUDED: 'Intentionally not populated (e.g. ECID under Web SDK — the SDK manages identity).'}
ORDER = [ST_COLLECTED, ST_NOOP, ST_ORPHAN, ST_NOTSENT, ST_NOTIMPL, ST_EXCLUDED]

# Slalom brand accents
BRAND_NAVY = '1F4E78'
HDR_FILL = PatternFill('solid', fgColor=BRAND_NAVY)
TIER_COLORS = {'P1': ('C00000', 'Decisions we need from the client'),
               'P2': ('BF8F00', 'Analytics-owner confirmations (no build change)'),
               'P3': ('1F6E43', 'Housekeeping to finalize the picture')}
SEV_FILL = {'High': 'FFC7CE', 'Medium': 'FFEB9C', 'Low': 'FCE4D6', 'Info': 'D9D9D9'}

thin = Side(style='thin', color='BFBFBF')
BORDER = Border(thin, thin, thin, thin)


def A(**k):
    return Font(name='Arial', **k)


def style_header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, c, h)
        cell.font = A(bold=True, color='FFFFFF')
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'


def write_rows(ws, rows, status_col):
    for r, row in enumerate(rows, start=2):
        st = row[status_col - 1]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(vertical='top', wrap_text=(c >= status_col))
            cell.border = BORDER
            if st in FILL:
                cell.fill = PatternFill('solid', fgColor=FILL[st])
                cell.font = A(size=10, color=FONTCOL.get(st, '000000'), bold=(c == status_col))
            else:
                cell.font = A(size=10)
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows) + 1}"


def var_status(kind, num, mapping, noop, overrides):
    """Return (status, feeding_refs) for a report-suite variable."""
    ov = (overrides.get(kind) or {}).get(str(num))
    refs = (mapping.get(kind) or {}).get(str(num)) or []
    if ov and ov.get('status'):
        return CANON.get(ov['status'], ov['status']), refs
    if refs:
        if any((r.get('element') in noop) for r in refs):
            return ST_NOOP, refs
        return ST_COLLECTED, refs
    return ST_NOTSENT, refs


def var_rows(kind, rsdict, mapping, noop, overrides):
    rows = []
    for num in sorted(rsdict, key=lambda x: int(x)):
        name = rsdict[num]
        st, refs = var_status(kind, num, mapping, noop, overrides)
        elems = '; '.join(sorted({r['element'] for r in refs if r.get('element')})) or '—'
        rules = '; '.join(sorted({r['rule'] for r in refs if r.get('rule')})) or '—'
        srcs = '; '.join(sorted({str(r.get('source', '')) for r in refs if r.get('source')})) or '—'
        ov = (overrides.get(kind) or {}).get(str(num)) or {}
        note = ov.get('note', '')
        label = {'eVars': 'eVar', 'props': 'prop', 'events': 'event', 'lists': 'list'}[kind]
        rows.append([f'{label}{num}', name, st, elems, rules, srcs, note])
    return rows


VALID_STATUS = set(CANON) | set(ORDER)


def validate(data):
    """Return a list of human-readable warnings about the input. Catches the
    dangerous silent failures: a mapping/override pointing at a variable that
    isn't defined in the report suite (a typo that would otherwise just vanish),
    unknown status values, and duplicate element names."""
    warns = []
    rs = data.get('reportSuite', {})
    mapping = data.get('mapping', {})
    overrides = data.get('statusOverrides', {})
    if not rs:
        warns.append("reportSuite is empty — there is nothing to reconcile against.")
    for kind in ('eVars', 'props', 'events', 'lists'):
        defined = set(rs.get(kind, {}))
        for k in (mapping.get(kind) or {}):
            if k not in defined:
                warns.append(f"mapping.{kind}['{k}'] is NOT a defined {kind[:-1]} in reportSuite — "
                             f"it will be silently dropped from the workbook (variable-number typo?).")
        for k, ov in (overrides.get(kind) or {}).items():
            if k not in defined:
                warns.append(f"statusOverrides.{kind}['{k}'] is NOT a defined {kind[:-1]} in reportSuite — ignored.")
            s = (ov or {}).get('status')
            if s and s not in VALID_STATUS:
                warns.append(f"statusOverrides.{kind}['{k}'].status='{s}' is not a known status "
                             f"(use one of: {', '.join(sorted(CANON))}).")
    names = [e.get('name') for e in data.get('elements', [])]
    dupes = sorted({n for n in names if names.count(n) > 1})
    if dupes:
        warns.append(f"duplicate names in elements[]: {dupes} — only the last row for each survives.")
    for e in data.get('elements', []):
        s = e.get('status')
        if s and s not in VALID_STATUS:
            warns.append(f"elements['{e.get('name')}'].status='{s}' is not a known status.")
    return warns


def main():
    argv = [a for a in sys.argv[1:] if not a.startswith('--')]
    strict = '--strict' in sys.argv
    if len(argv) != 2:
        print(__doc__)
        sys.exit(2)
    data = json.load(open(argv[0]))
    out = argv[1]

    warns = validate(data)
    if warns:
        sys.stderr.write(f"\n⚠  VALIDATION WARNINGS ({len(warns)}):\n")
        for w in warns:
            sys.stderr.write(f"   - {w}\n")
        sys.stderr.write("\n")
        if strict:
            sys.stderr.write("--strict set: aborting without writing the workbook.\n")
            sys.exit(1)

    meta = data.get('meta', {})
    rs = data.get('reportSuite', {})
    mapping = data.get('mapping', {})
    noop = set(data.get('noopElements', []))
    overrides = data.get('statusOverrides', {})
    elements = data.get('elements', [])
    issues = data.get('issues', [])
    recs = data.get('recommendations', [])

    wb = Workbook()

    # per-dimension rows (status in col 3), events have a slimmer layout
    erows = var_rows('eVars', rs.get('eVars', {}), mapping, noop, overrides)
    prows = var_rows('props', rs.get('props', {}), mapping, noop, overrides)
    vrows = var_rows('events', rs.get('events', {}), mapping, noop, overrides)
    lrows = var_rows('lists', rs.get('lists', {}), mapping, noop, overrides)

    # ---- eVars ----
    ws = wb.active
    ws.title = 'eVars'
    H = ['eVar #', 'Report Suite Name', 'Status', 'Data Element(s)', 'Sent By Rule(s)',
         'Data Element Source', 'Notes / Issue']
    style_header(ws, H, [8, 26, 26, 30, 34, 34, 46])
    write_rows(ws, erows, 3)

    # ---- Props ----
    ws = wb.create_sheet('Props')
    style_header(ws, ['Prop #'] + H[1:], [8, 26, 26, 30, 34, 34, 46])
    write_rows(ws, prows, 3)

    # ---- Events ----
    ws = wb.create_sheet('Events')
    HE = ['Event #', 'Report Suite Name', 'Status', 'Data Element(s)', 'Sent By Rule(s)', 'Notes / Issue']
    style_header(ws, HE, [8, 28, 26, 30, 40, 46])
    ev_rows = [[r[0], r[1], r[2], r[3], r[4], r[6]] for r in vrows]  # drop source col
    write_rows(ws, ev_rows, 3)

    # ---- List Vars (only if defined) ----
    if lrows:
        ws = wb.create_sheet('List Vars')
        HL = ['List Var', 'Report Suite Name', 'Status', 'Data Element(s)', 'Sent By Rule(s)', 'Notes / Issue']
        style_header(ws, HL, [10, 24, 26, 30, 40, 46])
        lv_rows = [[r[0], r[1], r[2], r[3], r[4], r[6]] for r in lrows]
        write_rows(ws, lv_rows, 3)

    # ---- Data Elements (element -> variable) ----
    ws = wb.create_sheet('Data Elements')
    HD = ['Data Element', 'Feeds Variable(s)', 'Status', 'Notes']
    style_header(ws, HD, [30, 40, 26, 60])
    de_rows = []
    for e in sorted(elements, key=lambda x: x['name']):
        feeds = e.get('feeds', [])
        st = e.get('status')
        if st:
            st = CANON.get(st, st)
        elif e['name'] in noop:
            st = ST_NOOP
        elif feeds:
            st = ST_COLLECTED
        else:
            st = ST_ORPHAN
        de_rows.append([e['name'], '; '.join(feeds) if feeds else '—', st, e.get('note', '')])
    write_rows(ws, de_rows, 3)

    # ---- Data Issues ----
    ws = wb.create_sheet('Data Issues')
    HI = ['#', 'Issue', 'Severity', 'Affected Variables / Elements', 'Description', 'Recommendation']
    style_header(ws, HI, [5, 30, 12, 28, 52, 46])
    for i, iss in enumerate(issues, start=1):
        vals = [i, iss.get('issue', ''), iss.get('severity', 'Info'), iss.get('affected', ''),
                iss.get('description', ''), iss.get('recommendation', '')]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(i + 1, c, v)
            cell.font = A(size=10, bold=(c == 3))
            cell.alignment = Alignment(vertical='top', wrap_text=(c in (2, 4, 5, 6)))
            cell.border = BORDER
        sev = iss.get('severity', 'Info')
        ws.cell(i + 1, 3).fill = PatternFill('solid', fgColor=SEV_FILL.get(sev, 'FFFFFF'))
    if issues:
        ws.auto_filter.ref = f"A1:F{len(issues) + 1}"

    # ---- Recommendations ----
    ws = wb.create_sheet('Recommendations', 1)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 54
    ws.column_dimensions['C'].width = 74
    ws.column_dimensions['D'].width = 16
    ws['B1'] = 'Recommendations & Next Steps'
    ws['B1'].font = A(size=16, bold=True, color=BRAND_NAVY)
    sub = meta.get('bottomLine', 'Items below are decisions and clean-ups the reconciliation surfaced. '
                                 'Detail per item is on the Data Issues tab.')
    ws['B2'] = sub
    ws['B2'].font = A(size=10, italic=True, color='595959')
    ws.merge_cells('B2:D2')
    ws['B2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 42
    r = 4
    for tier in ('P1', 'P2', 'P3'):
        items = [x for x in recs if x.get('tier') == tier]
        if not items:
            continue
        color, label = TIER_COLORS[tier]
        for cc in (2, 3, 4):
            cell = ws.cell(r, cc)
            cell.fill = PatternFill('solid', fgColor=color)
            cell.font = A(size=11, bold=True, color='FFFFFF')
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center')
        ws.cell(r, 2, label)
        ws.row_dimensions[r].height = 22
        r += 1
        for cc, h in [(2, 'Recommendation'), (3, 'Why it matters'), (4, 'When')]:
            ws.cell(r, cc, h).font = A(size=9, bold=True, color='808080')
        ws.row_dimensions[r].height = 15
        r += 1
        for it in items:
            a = ws.cell(r, 2, it.get('recommendation', ''))
            a.font = A(size=10, bold=True)
            a.alignment = Alignment(wrap_text=True, vertical='top')
            b = ws.cell(r, 3, it.get('why', ''))
            b.font = A(size=10)
            b.alignment = Alignment(wrap_text=True, vertical='top')
            d = ws.cell(r, 4, it.get('when', ''))
            d.font = A(size=10, italic=True, color=color)
            d.alignment = Alignment(wrap_text=True, vertical='top')
            for cc in (2, 3, 4):
                ws.cell(r, cc).border = Border(bottom=Side(style='thin', color='E0E0E0'))
            ws.row_dimensions[r].height = 34
            r += 1
        r += 1

    # ---- Overview (first) — combined legend + counts ----
    ws = wb.create_sheet('Overview', 0)
    ws.sheet_view.showGridLines = False
    ws['A1'] = f"{meta.get('property', 'Property')} — Data Collection Reconciliation"
    ws['A1'].font = A(size=16, bold=True, color=BRAND_NAVY)
    ws['A2'] = meta.get('subtitle', "Report suite variables mapped to the property's data elements — "
                                    "what is collecting data, what is not wired up, and open data issues.")
    ws['A2'].font = A(size=10, italic=True, color='595959')
    cap = f"Captured {meta.get('captured', '')}"
    if meta.get('reportSuite'):
        cap += f" · report suite: {meta['reportSuite']}"
    if meta.get('note'):
        cap += f" · {meta['note']}"
    ws['A3'] = cap
    ws['A3'].font = A(size=9, color='808080')
    ws['A5'] = 'Status legend & counts'
    ws['A5'].font = A(size=12, bold=True)
    ws['A6'] = 'Counts show how many of each variable / element carry that status. Full detail is on the tabs below.'
    ws['A6'].font = A(size=9, italic=True, color='808080')

    def cnt(rows, st):
        return sum(1 for row in rows if row[2] == st)

    def dcnt(st):
        return sum(1 for row in de_rows if row[2] == st)

    hdr_row = 8
    COLS = ['Status', 'eVars', 'Props', 'Events', 'Lists', 'Data El.', 'What it means']
    CW = [30, 8, 8, 8, 8, 9, 88]
    for c, (h, w) in enumerate(zip(COLS, CW), start=1):
        cell = ws.cell(hdr_row, c, h)
        cell.font = A(bold=True, color='FFFFFF')
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal=('left' if c in (1, 7) else 'center'), vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[hdr_row].height = 20
    r = hdr_row + 1
    for st in ORDER:
        vals = [st, cnt(erows, st), cnt(prows, st), cnt(vrows, st), cnt(lrows, st), dcnt(st), DESC[st]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.fill = PatternFill('solid', fgColor=FILL[st])
            cell.font = A(size=10, color=FONTCOL[st], bold=(c == 1))
            cell.alignment = Alignment(horizontal=('left' if c in (1, 7) else 'center'),
                                       vertical='center', wrap_text=(c == 7))
            cell.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1
    tot = ['Total (defined)', len(erows), len(prows), len(vrows), len(lrows), len(de_rows), '']
    for c, v in enumerate(tot, start=1):
        cell = ws.cell(r, c, v)
        cell.font = A(bold=True)
        cell.alignment = Alignment(horizontal=('left' if c in (1, 7) else 'center'), vertical='center')
        cell.border = BORDER
        cell.fill = PatternFill('solid', fgColor='F2F2F2')
    if meta.get('bottomLine'):
        r += 2
        ws.cell(r, 1, meta['bottomLine']).font = A(size=10, italic=True, color=BRAND_NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 30

    # enforce tab order (Overview, Recommendations, then the detail tabs)
    order = ['Overview', 'Recommendations', 'eVars', 'Props', 'Events', 'List Vars',
             'Data Elements', 'Data Issues']
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    wb.save(out)
    print(f"saved {out}")
    print(f"  tabs: {wb.sheetnames}")

    # Auto-reconcile: print the status breakdown per dimension so a wrong count
    # is visible in the build log instead of needing a manual eyeball pass.
    from collections import Counter
    print("  reconcile (status breakdown = total):")
    for name, rows in [('eVars', erows), ('props', prows), ('events', vrows),
                       ('lists', lrows), ('elements', de_rows)]:
        if not rows:
            continue
        c = Counter(r[2] for r in rows)
        short = {ST_COLLECTED: 'collected', ST_NOOP: 'no-op', ST_ORPHAN: 'orphan',
                 ST_NOTSENT: 'never-sent', ST_NOTIMPL: 'not-built', ST_EXCLUDED: 'excluded'}
        parts = ' + '.join(f"{v} {short.get(k, k)}" for k, v in c.items())
        print(f"    {name:9}: {parts} = {sum(c.values())}")
    if warns:
        print(f"  NOTE: {len(warns)} validation warning(s) above — review before delivering.")


if __name__ == '__main__':
    main()
