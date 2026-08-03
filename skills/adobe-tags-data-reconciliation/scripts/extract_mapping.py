#!/usr/bin/env python3
"""
DRAFT extractor: read an adobe-tags-audit-builder audit workbook (.xlsx) and
emit a first-pass variable -> data-element mapping for the reconciliation.

This does the tedious, error-prone half of SKILL.md Phase 2 mechanically — it
scans every rule Action for analytics-variable assignments (both classic
AppMeasurement `s.eVar5 = "%Page Name%"` and Web SDK `data.__adobe.analytics`
`"eVar5": "%Page Name%"` forms) and inverts them into a draft mapping. Claude
then EDITS that draft (fills report-suite variable names, resolves the
unmatched/low-confidence rows, moves no-ops/orphans/excluded items) rather than
hand-building the whole mapping from scratch.

    python3 extract_mapping.py audit-workbook.xlsx draft-mapping.json

It reads three tabs of the audit workbook:
  - "Rule Detail"          — Action rows: the variable assignments (the core)
  - "Data Elements"        — the authoritative element-name list + what each resolves to
  - "Data Element Detail"  — element code, scanned for getVar() cross-references

What it CANNOT know (and why the output is a draft, not the answer):
  - Report-suite variable NAMES (eVar5 = "Internal Campaign") live in the
    client's Report Suite Manager, not the audit workbook. Fields are left blank.
  - Whether an element is a no-op (`return ""` / `console.log()`), an intentional
    exclusion (ECID), or a genuine orphan — that's Phase 3 judgment.
  - A value set by an expression/literal rather than a data element — those go to
    `unmatched[]` for a human to classify, never silently into the mapping.

Output shape (a superset of build_reconciliation_workbook.py's `mapping`, so the
resolved parts paste straight in):

{
  "_draft": true,
  "_note": "...how to use this draft...",
  "mapping": {                                  // forward: variable -> element(s)
    "eVars":  {"5": [{"element": "Internal Campaign", "rule": "Global Page Load Rule",
                      "source": "custom code", "confidence": "high"}]},
    "props":  {...}, "events": {...}, "lists": {...}
  },
  "topLevelFields": {                           // s.pageName / s.campaign / ... -> element
    "pageName": [{"element": "Page Name", "rule": "Global Page Load Rule"}]
  },
  "elementsDraft": [                            // reverse: element -> what it feeds
    {"name": "Site: Domain", "feeds": ["eVar1"]},
    {"name": "Orphan Candidate", "feeds": []}   // seen in no action & no getVar
  ],
  "unmatched": [                               // analytics fields we couldn't resolve to an element
    {"rule": "...", "field": "eVar12", "rawValue": "'D=g'",
     "reason": "inline literal / expression — no data element"}
  ],
  "unknownElements": ["Ghost Element"],        // %refs% not found on the Data Elements tab
  "stats": {...}
}
"""
import json
import re
import sys
from collections import defaultdict, Counter
from openpyxl import load_workbook

# analytics fields we recognize on the left of an assignment
NUMBERED = r'eVar\d+|prop\d+|list\d+'
TOPLEVEL = (r'pageName|channel|campaign|referrer|pageURL|server|purchaseID|'
            r'transactionID|products|state|zip|purchaseid')
FIELD = rf'(?:{NUMBERED}|events|{TOPLEVEL})'

# A field assignment in either JSON ("eVar5": "%X%") or code (s.eVar5 = "%X%").
# The field token must start on a boundary so "linkTrackEvents" can't match "events".
ASSIGN = re.compile(
    r'(?:^|[\s{,.;(])(?:s\.)?["\']?(' + FIELD + r')["\']?\s*(?<![=<>!])[:=](?!=)\s*(.+?)(?=[,;}\n]|$)',
    re.I)
# Classic Analytics-extension structured form: {"name":"eVar5", ... "value":"%X%"}
NAME_TOKEN = re.compile(r'"name"\s*:\s*"(' + NUMBERED + r'|event\d+)"', re.I)
VALUE_NEAR = re.compile(r'"value"\s*:\s*"([^"]*)"', re.I)
TYPE_NEAR = re.compile(r'"type"\s*:\s*"(\w+)"', re.I)

PCT = re.compile(r'%([^%]+)%')                                  # %Data Element%
GETVAR = re.compile(r'getVar\(\s*["\']([^"\']+)["\']', re.I)    # _satellite.getVar("X")
EVENT_TOKEN = re.compile(r'(event\d+)', re.I)
ALIAS = re.compile(r'^(eVar\d+|prop\d+)$', re.I)

KIND = {'evar': 'eVars', 'prop': 'props', 'list': 'lists', 'event': 'events'}


def sheet(wb, *wanted):
    """Case-insensitive sheet lookup; returns the worksheet or None."""
    low = {ws.title.strip().lower(): ws for ws in wb.worksheets}
    for w in wanted:
        if w.lower() in low:
            return low[w.lower()]
    return None


def header_index(ws):
    """Map lowercased header name -> 0-based column index from row 1."""
    idx = {}
    for c, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if cell:
            idx[str(cell).strip().lower()] = c
    return idx


def split_field(tok):
    m = re.match(r'(eVar|prop|list|event)(\d+)$', tok, re.I)
    if m:
        return KIND[m.group(1).lower()], m.group(2)
    return 'top', tok


def resolve_element(rhs):
    """Return (element_name, how) from an assignment RHS, or (None, None)."""
    rhs = (rhs or '').strip()
    m = PCT.search(rhs)
    if m:
        return m.group(1).strip(), 'percent'
    m = GETVAR.search(rhs)
    if m:
        return m.group(1).strip(), 'getvar'
    return None, None


def scan_action(detail, rule, found, unmatched, seen):
    """Scan one Action's Full Detail; append discovered feeds to `found` and
    unresolved analytics fields to `unmatched`. `seen` dedupes (kind,num,elem,rule)."""
    if not detail:
        return
    # 1) direct key/value + code assignments. The 'events' KEY (plural, no number)
    # is special-cased BEFORE split_field, which only understands eventN tokens.
    for m in ASSIGN.finditer(detail):
        field, rhs = m.group(1), m.group(2)
        if field.lower() == 'events':
            for ev in EVENT_TOKEN.findall(rhs):
                _add(found, seen, 'events', ev[5:], resolve_element(rhs)[0] or '(no value)',
                     rule, 'custom code', 'high')
            continue
        kind, num = split_field(field)
        elem, how = resolve_element(rhs)
        _classify(kind, num, field, elem, how, rhs, rule, found, unmatched, seen)
    # 2) classic Analytics-extension structured {"name":"eVar5",..."value":"%X%"}.
    # Bound each object's window at the NEXT "name" token so a valued event can't
    # bleed its value onto the preceding (value-less) event.
    names = list(NAME_TOKEN.finditer(detail))
    for i, m in enumerate(names):
        tok = m.group(1)
        stop = names[i + 1].start() if i + 1 < len(names) else m.start() + 260
        win = detail[m.start():min(stop, m.start() + 260)]
        kind, num = split_field(tok)
        if kind == 'events':
            v = VALUE_NEAR.search(win)
            _add(found, seen, 'events', num, (resolve_element(v.group(1))[0] if v else None) or '(no value)',
                 rule, 'set variables', 'high')
            continue
        v = VALUE_NEAR.search(win)
        t = TYPE_NEAR.search(win)
        if t and t.group(1).lower() == 'alias' and v:
            _add(found, seen, kind, num, f'(alias of {v.group(1)})', rule, 'alias', 'low')
            continue
        elem, how = resolve_element(v.group(1) if v else '')
        _classify(kind, num, tok, elem, how, v.group(1) if v else '', rule, found, unmatched, seen)


def _classify(kind, num, field, elem, how, rhs, rule, found, unmatched, seen):
    # Leave source blank; the enrich pass fills it from the element's own
    # definition (which usefully reveals no-ops like `return "";`).
    if elem:
        _add(found, seen, kind, num, elem, rule, '', 'high')
    else:
        unmatched.append({'rule': rule, 'field': field, 'rawValue': (rhs or '').strip()[:120],
                          'reason': 'inline literal / expression — no data element reference'})


def _add(found, seen, kind, num, elem, rule, source, confidence):
    key = (kind, num, elem, rule)
    if key in seen:
        return
    seen.add(key)
    found[kind].setdefault(num, []).append(
        {'element': elem, 'rule': rule, 'source': source, 'confidence': confidence})


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    wb = load_workbook(sys.argv[1], read_only=True, data_only=True)

    de_ws = sheet(wb, 'Data Elements')
    det_ws = sheet(wb, 'Data Element Detail')
    rd_ws = sheet(wb, 'Rule Detail')
    if rd_ws is None:
        sys.stderr.write("ERROR: no 'Rule Detail' tab found — is this an adobe-tags-audit-builder workbook?\n")
        sys.exit(1)

    # element names + a short 'resolves to' for source hints
    known, resolves = set(), {}
    if de_ws is not None:
        di = header_index(de_ws)
        nc = di.get('data element', 0)
        rc = di.get('resolves to')
        for row in de_ws.iter_rows(min_row=2, values_only=True):
            if row and row[nc]:
                name = str(row[nc]).strip()
                known.add(name)
                if rc is not None and row[rc]:
                    resolves[name] = str(row[rc]).strip()

    # element -> element getVar references (so a "campaign string" element that
    # reads UTM elements isn't mis-flagged as an orphan)
    getvar_used = set()
    if det_ws is not None:
        hi = header_index(det_ws)
        dc = hi.get('full detail', 1)
        for row in det_ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > dc and row[dc]:
                for t in GETVAR.findall(str(row[dc])):
                    getvar_used.add(t.strip())

    found = defaultdict(dict)   # kind -> {num -> [rows]}; kind includes 'top'
    unmatched, seen = [], set()
    rules_scanned, actions_scanned = set(), 0

    hi = header_index(rd_ws)
    c_rule = hi.get('rule name', 1)
    c_type = hi.get('item type', 2)
    c_sum = hi.get('item summary', 3)
    c_det = hi.get('full detail', 4)
    for row in rd_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        itype = str(row[c_type]).strip().lower() if len(row) > c_type and row[c_type] else ''
        if 'action' not in itype:
            continue
        rule = str(row[c_rule]).strip() if len(row) > c_rule and row[c_rule] else '(unnamed rule)'
        rules_scanned.add(rule)
        actions_scanned += 1
        detail = row[c_det] if len(row) > c_det and row[c_det] else ''
        if not detail and len(row) > c_sum and row[c_sum]:
            detail = row[c_sum]
        scan_action(str(detail), rule, found, unmatched, seen)

    # enrich source hint from the element's 'resolves to', and collect unknown refs
    unknown = set()
    for kind in list(found):
        for num, rows in found[kind].items():
            for rrow in rows:
                el = rrow['element']
                if el.startswith('(') or kind == 'events':
                    continue
                if el not in known:
                    unknown.add(el)
                if not rrow.get('source') and el in resolves:
                    rrow['source'] = resolves[el][:60]
                elif not rrow.get('source'):
                    rrow['source'] = 'custom code'

    # reverse view: element -> feeds[]
    feeds = defaultdict(list)
    for kind in ('eVars', 'props', 'lists', 'events'):
        for num, rows in found.get(kind, {}).items():
            label = {'eVars': 'eVar', 'props': 'prop', 'lists': 'list', 'events': 'event'}[kind] + num
            for rrow in rows:
                if not rrow['element'].startswith('('):
                    feeds[rrow['element']].append(label)
    for field, rows in found.get('top', {}).items():
        for rrow in rows:
            if not rrow['element'].startswith('('):
                feeds[rrow['element']].append(f's.{field}')
    for el in getvar_used:
        if el in known:
            feeds[el].append('getVar (another element)')
    elements_draft = []
    for name in sorted(known):
        elements_draft.append({'name': name, 'feeds': sorted(set(feeds.get(name, [])))})
    # elements referenced by %..%/getVar but absent from the Data Elements tab
    for name in sorted(set(feeds) - known):
        elements_draft.append({'name': name, 'feeds': sorted(set(feeds[name])),
                               '_note': 'referenced but NOT on the Data Elements tab — verify name'})

    top = {f: rows for f, rows in found.get('top', {}).items()}
    out = {
        '_draft': True,
        '_note': ("DRAFT auto-extraction from the audit workbook — REVIEW every row. "
                  "(1) Fill each variable's report-suite NAME from the client's Report Suite "
                  "Manager export (not in the audit workbook). (2) Resolve unmatched[] — each is "
                  "an analytics field set by a literal/expression, not a data element. "
                  "(3) Move confirmed no-ops to noopElements, orphans/exclusions to statusOverrides. "
                  "(4) 'confidence:low' rows (aliases) and unknownElements[] need a second look. "
                  "'source' is a best-effort guess from the element's definition. "
                  "topLevelFields feed elements[].feeds (e.g. s.pageName), not report-suite variables."),
        'mapping': {k: dict(sorted(found.get(k, {}).items(), key=lambda x: int(x[0])))
                    for k in ('eVars', 'props', 'events', 'lists')},
        'topLevelFields': top,
        'elementsDraft': elements_draft,
        'unmatched': unmatched,
        'unknownElements': sorted(unknown),
        'stats': {
            'rulesScanned': len(rules_scanned),
            'actionsScanned': actions_scanned,
            'eVars': len(found.get('eVars', {})),
            'props': len(found.get('props', {})),
            'events': len(found.get('events', {})),
            'lists': len(found.get('lists', {})),
            'topLevel': len(top),
            'unmatched': len(unmatched),
            'unknownElements': len(unknown),
            'dataElements': len(known),
        },
    }
    json.dump(out, open(sys.argv[2], 'w'), indent=2)

    s = out['stats']
    print(f"wrote {sys.argv[2]}")
    print(f"  scanned {s['actionsScanned']} actions across {s['rulesScanned']} rules")
    print(f"  mapped: {s['eVars']} eVars · {s['props']} props · {s['events']} events · "
          f"{s['lists']} lists · {s['topLevel']} top-level fields")
    print(f"  {s['unmatched']} unmatched analytics field(s), {s['unknownElements']} unknown element ref(s)")
    if unknown:
        print(f"    unknown refs: {', '.join(sorted(unknown)[:8])}{' …' if len(unknown) > 8 else ''}")
    print("  NOTE: this is a DRAFT — edit it (fill variable names, resolve unmatched[]) before "
          "feeding build_reconciliation_workbook.py.")


if __name__ == '__main__':
    main()
